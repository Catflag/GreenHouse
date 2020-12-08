import Adafruit_DHT
import time
import RPi.GPIO as GPIO
from openpyxl import load_workbook
from openpyxl import Workbook
sensor=Adafruit_DHT.DHT11
GPIO.setwarnings(False)
GPIO.setmode(GPIO.BOARD)
GPIO.setup(12,GPIO.OUT)

# Set GPIO sensor is connected to 17


humidity, temperature = Adafruit_DHT.read_retry(sensor, 17)
 
if (humidity is not None and temperature is not None):
  print('Temp={0:0.1f}*C  Humidity={1:0.1f}%'.format(temperature, humidity))
  wb=load_workbook('Temp_Humidity.xlsx')
  ws1=wb.active
  ws1['A2'].value='Temp_Humi'
  ws1['B2'].value='temperture'
  ws1['C2'].value='humidity'
  wb.save('Temp_Humidity.xlsx')
else:
  print('Failed to get reading. Try again!')

#Define Fan control by Temperature code
if temperature <= 35:
  GPIO.output(12,GPIO.HIGH)
  print("Exhaust OFF")
  wb=load_workbook('Temp_Humidity.xlsx')
  ws1=wb.active
  ws1['A3'].value='Exhaust_Fan'
  ws1['B3'].value='OFF'
  wb.save('Temp_Humidity.xlsx')
  time.sleep(5)
else:
  GPIO.output(12,GPIO.LOW)
  print("Exhaust ON")
  wb=load_workbook('Temp_Humidity.xlsx')
  ws1=wb.active
  ws1['A3'].value='Exhaust_Fan'
  ws1['B3'].value='ON'
  wb.save('Temp_Humidity.xlsx')
  time.sleep(5)
  
