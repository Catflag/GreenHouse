import sys
import RPi.GPIO as GPIO # GPIO library we need to use the GPIO pins
import time # time library for sleep function
from openpyxl import load_workbook
from openpyxl import Workbook


StepPinForward=32
StepPinBackward=36
GPIO.setmode(GPIO.BOARD)
GPIO.setup(StepPinForward, GPIO.OUT)
GPIO.setup(StepPinBackward, GPIO.OUT)

channel = 29
GPIO.setup(channel, GPIO.IN)

def soilsensor(channel):  
    if GPIO.input(channel):
        print ("Start to Pump!")
        pumpforward(3)
    else:       
        print ("Water Is Detected")
        GPIO.output(StepPinForward, GPIO.LOW)
        wb=load_workbook('Soil_Water_control.xlsx')
        ws1=wb.active
        ws1['A2'].value='Soil_Water'
        ws1['B2'].value='Wa_Detect'
        wb.save('Soil_Water_control.xlsx')       


def pumpforward(x):
    GPIO.output(StepPinForward, GPIO.HIGH)
    print ("Water Pump Start")
    wb=load_workbook('Soil_Water_control.xlsx')
    ws1=wb.active
    ws1['A3'].value='Water_Pump'
    ws1['B3'].value='Start'
    wb.save('Soil_Water_control.xlsx')           
    time.sleep(5)
    GPIO.output(StepPinForward, GPIO.LOW)


def pumpbackward(x):
    GPIO.output(StepPinBackward, GPIO.HIGH)
    print ("Water Pump Stoped")
    wb=load_workbook('Soil_Water_control.xlsx')
    ws1=wb.active
    ws1['A3'].value='Water_Pump'
    ws1['B3'].value='Stop'
    wb.save('Soil_Water_control.xlsx')           
    time.sleep(120)
    GPIO.output(StepPinBackward, GPIO.LOW)
               
GPIO.add_event_detect(channel, GPIO.BOTH, bouncetime=300)
GPIO.add_event_callback(channel, soilsensor)

# Loop Continue ON
while True:
    time.sleep(0.1)
